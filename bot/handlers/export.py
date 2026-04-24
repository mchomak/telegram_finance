import io
import logging
from calendar import monthrange
from datetime import date, timedelta

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from bot.db.models import Category, Expense, SharedExpense
from bot.utils.keyboards import export_period_keyboard

logger = logging.getLogger(__name__)
router = Router()

_MONTHS_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BOLD_FONT = Font(name="Calibri", bold=True, size=11)


# ── Period helpers ─────────────────────────────────────────────────────────────

def _period_bounds(period: str) -> tuple[date | None, date | None, str]:
    today = date.today()
    if period == "current_month":
        date_from = today.replace(day=1)
        date_to = today.replace(day=monthrange(today.year, today.month)[1])
        return date_from, date_to, f"расходы_{_MONTHS_RU[today.month]}_{today.year}.xlsx"
    if period == "last_month":
        last = today.replace(day=1) - timedelta(days=1)
        return last.replace(day=1), last, f"расходы_{_MONTHS_RU[last.month]}_{last.year}.xlsx"
    return None, None, "расходы_все_время.xlsx"


# ── DB helpers ─────────────────────────────────────────────────────────────────

async def _fetch_export_data(
    session: AsyncSession,
    user_id: int,
    date_from: date | None,
    date_to: date | None,
) -> tuple[list[tuple[Expense, Category | None]], dict[int, list[SharedExpense]], list[Category]]:
    cats = (
        await session.execute(
            select(Category)
            .where(Category.user_id == user_id, Category.is_active.is_(True))
            .order_by(Category.id)
        )
    ).scalars().all()

    q = (
        select(Expense, Category)
        .outerjoin(Category, Expense.category_id == Category.id)
        .where(Expense.user_id == user_id)
    )
    if date_from:
        q = q.where(Expense.expense_date >= date_from)
    if date_to:
        q = q.where(Expense.expense_date <= date_to)
    q = q.order_by(Expense.expense_date.desc(), Expense.created_at.desc())

    expense_rows = [(e, c) for e, c in (await session.execute(q)).all()]

    shared_map: dict[int, list[SharedExpense]] = {}
    if expense_rows:
        expense_ids = [e.id for e, _ in expense_rows]
        for s in (
            await session.execute(
                select(SharedExpense).where(SharedExpense.expense_id.in_(expense_ids))
            )
        ).scalars().all():
            shared_map.setdefault(s.expense_id, []).append(s)

    return expense_rows, shared_map, list(cats)


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _cat_label(cat: Category | None) -> str:
    if not cat:
        return "Без категории"
    return f"{cat.emoji} {cat.name}" if cat.emoji else cat.name


def _fmt_participants(shared: list[SharedExpense]) -> str:
    parts = []
    for s in shared:
        if s.amount_owed is not None:
            val = float(s.amount_owed)
            amt_str = str(int(val)) if val == int(val) else f"{val:.2f}"
            parts.append(f"{s.participant_name} ({amt_str}₽)")
        elif s.item_description:
            parts.append(f"{s.participant_name} ({s.item_description})")
        else:
            parts.append(s.participant_name)
    return ", ".join(parts)


def _fmt_return_status(shared: list[SharedExpense]) -> str:
    if not shared:
        return ""
    returned = sum(1 for s in shared if s.is_returned)
    if returned == len(shared):
        return "получен"
    if returned > 0:
        return "частично"
    return "ожидается"


# ── Excel builders ─────────────────────────────────────────────────────────────

def _build_summary_sheet(
    ws,
    expense_rows: list[tuple[Expense, Category | None]],
    categories: list[Category],
) -> None:
    # Ordered category labels: active categories first, then any extra from expenses
    cat_order: dict[str, int] = {_cat_label(c): i for i, c in enumerate(categories)}
    for _, cat in expense_rows:
        lbl = _cat_label(cat)
        if lbl not in cat_order:
            cat_order[lbl] = len(cat_order)
    all_cats = sorted(cat_order, key=cat_order.__getitem__)

    # Build day → category → total
    day_totals: dict[date, dict[str, float]] = {}
    for expense, cat in expense_rows:
        if expense.expense_date is None or expense.amount is None:
            continue
        d = expense.expense_date
        lbl = _cat_label(cat)
        day_totals.setdefault(d, {})
        day_totals[d][lbl] = day_totals[d].get(lbl, 0.0) + float(expense.amount)

    if not day_totals:
        ws.append(["Нет данных за выбранный период"])
        return

    sorted_days = sorted(day_totals)
    headers = ["Дата"] + all_cats + ["Итого"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row, d in enumerate(sorted_days, 2):
        row_data = day_totals[d]
        ws.cell(row=row, column=1, value=d.strftime("%d.%m.%Y"))
        row_sum = 0.0
        for col, lbl in enumerate(all_cats, 2):
            val = row_data.get(lbl, 0.0)
            if val:
                ws.cell(row=row, column=col, value=round(val, 2))
                row_sum += val
        ws.cell(row=row, column=len(all_cats) + 2, value=round(row_sum, 2)).font = _BOLD_FONT

    total_row = len(sorted_days) + 2
    ws.cell(row=total_row, column=1, value="Итого").font = _BOLD_FONT
    grand = 0.0
    for col, lbl in enumerate(all_cats, 2):
        s = sum(day_totals[d].get(lbl, 0.0) for d in sorted_days)
        if s:
            ws.cell(row=total_row, column=col, value=round(s, 2)).font = _BOLD_FONT
        grand += s
    ws.cell(row=total_row, column=len(all_cats) + 2, value=round(grand, 2)).font = _BOLD_FONT

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 4)
    ws.freeze_panes = "A2"


def _build_history_sheet(
    ws,
    expense_rows: list[tuple[Expense, Category | None]],
    shared_map: dict[int, list[SharedExpense]],
) -> None:
    headers = [
        "Дата", "Время", "Расшифровка", "Описание",
        "Категория", "Сумма", "Участники", "Статус возврата",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row, (expense, cat) in enumerate(expense_rows, 2):
        shared = shared_map.get(expense.id, [])
        ws.cell(
            row=row, column=1,
            value=expense.expense_date.strftime("%d.%m.%Y") if expense.expense_date else "",
        )
        ws.cell(
            row=row, column=2,
            value=expense.created_at.strftime("%H:%M") if expense.created_at else "",
        )
        transcription_cell = ws.cell(row=row, column=3, value=expense.transcription or "")
        transcription_cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4, value=expense.note or "")
        ws.cell(row=row, column=5, value=_cat_label(cat) if cat else "")
        if expense.amount is not None:
            ws.cell(row=row, column=6, value=float(expense.amount))
        ws.cell(row=row, column=7, value=_fmt_participants(shared))
        ws.cell(row=row, column=8, value=_fmt_return_status(shared))

    for col, width in enumerate([12, 8, 40, 25, 18, 10, 35, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def _generate_excel(
    expense_rows: list[tuple[Expense, Category | None]],
    shared_map: dict[int, list[SharedExpense]],
    categories: list[Category],
) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сводка"
    _build_summary_sheet(ws1, expense_rows, categories)

    ws2 = wb.create_sheet("История")
    _build_history_sheet(ws2, expense_rows, shared_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Handlers ──────────────────────────────────────────────────────────────────

@router.message(F.text == "📤 Экспорт")
async def cmd_export(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("Выберите период для экспорта:", reply_markup=export_period_keyboard())


async def _handle_export_period(
    callback: CallbackQuery,
    session: AsyncSession,
    period: str,
) -> None:
    date_from, date_to, filename = _period_bounds(period)
    await callback.answer("Генерирую файл…")

    expense_rows, shared_map, categories = await _fetch_export_data(
        session, callback.from_user.id, date_from, date_to
    )

    if not expense_rows:
        await callback.message.edit_text("За выбранный период расходов не найдено.")
        return

    xlsx_bytes = _generate_excel(expense_rows, shared_map, categories)
    await callback.message.answer_document(
        document=BufferedInputFile(xlsx_bytes, filename=filename),
    )
    try:
        await callback.message.delete()
    except Exception:
        pass


@router.callback_query(F.data == "export:current_month")
async def cb_export_current(callback: CallbackQuery, session: AsyncSession) -> None:
    await _handle_export_period(callback, session, "current_month")


@router.callback_query(F.data == "export:last_month")
async def cb_export_last(callback: CallbackQuery, session: AsyncSession) -> None:
    await _handle_export_period(callback, session, "last_month")


@router.callback_query(F.data == "export:all_time")
async def cb_export_all(callback: CallbackQuery, session: AsyncSession) -> None:
    await _handle_export_period(callback, session, "all_time")
